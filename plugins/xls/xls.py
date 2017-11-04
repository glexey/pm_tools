import os
import re
import warnings
import openpyxl
import xml.etree.cElementTree as ET

# Turn off some annoying warnings from openpyxl
warnings.filterwarnings("ignore", "Cannot parse header or footer so it will be ignored", module='openpyxl')
warnings.filterwarnings("ignore", "Discarded range with reserved name", module='openpyxl')
warnings.filterwarnings("ignore", "Using a range string is deprecated. Use ws\[range_string\]", module='openpyxl')
warnings.filterwarnings("ignore", "Conditional Formatting extension is not supported and will be removed", module='openpyxl')
warnings.filterwarnings("ignore", "Unknown extension is not supported and will be removed", module='openpyxl')
warnings.filterwarnings("ignore", "Data Validation extension is not supported and will be removed", module='openpyxl')

class XlsPlugin(object):

    def __init__(self, preprocessor):
        self.pp = preprocessor
        self.token = "xls"
        self.pp.register_plugin(self)
        self.excel = {}
        self.comment_gid = 0 # global comment ID for relative refs in this file.

    def process(self, code, fname, sheet="", title=None, range="", div_style=None, header=True):
        """
        Open up a specific tab of an xls file and dump it to HTML

        ```xls("fname", "optional sheet name", "Optional Title" [, range="cell_range"][, header=False])

        """
        if title is None:
            # construct default title
            atitle = []
            if sheet != '': atitle.append(sheet)
            if range != '': atitle.append(range)
            if atitle == []: atitle.append(os.path.splitext(os.path.basename(fname))[0])
            title = '_'.join(atitle)

        # Pull from the web if the user wants it
        dstfile = fname
        if (fname.startswith("http")):
            try:
                dstfile = save_url(fname, os.path.join(self.pp.dirs[-1], self.auto))
            except:
                # deal with a bad URL
                return "[%s <mark>_file not found_</mark>](%s)" % (title, fname)

        xlsfile, relative_path = self.pp.get_asset(dstfile, True)

        # For web resource, provide a link to the original
        if (fname.startswith("http")):
            link = fname
        else:
            link = relative_path

        # Read the xls range into a rich table object
        xltable = self.readxls(xlsfile, sheet, range, link, header=header)

        xltable.title = title
        return xltable.html()

    def readxls(self, xlsfile, sheet, _range="", relative_path=None, header=True):
        """
        Import a range from xls file into local workbook / cell objects for displaying in our doc
        """

        if (self.excel.has_key(xlsfile)):
            # If this file was already loaded, re-use it
            wb, wb_formula = self.excel[xlsfile]
        else:
            # Load both data only (for raw cell data) and the formulas
            # Workbook object will merge the two.
            wb = openpyxl.load_workbook(filename=xlsfile, data_only=True)
            wb_formula = openpyxl.load_workbook(filename=xlsfile)
            self.excel[xlsfile] = [wb, wb_formula]

            # Work-around for openpyxl not handling theme colors
            theme_root = ET.fromstring(wb.loaded_theme)
            ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
            scheme = theme_root.find('.//a:clrScheme', ns) # use first found color scheme
            wb.theme_colors = tc = [x.attrib.get('lastClr', x.attrib.get('val', None)) for x in scheme.findall('*/*')]
            # http://stackoverflow.com/questions/2760976/theme-confusion-in-spreadsheetml
            for i in [0, 2]: tc[i], tc[i+1] = tc[i+1], tc[i]

        return Worksheet(wb, wb_formula, sheet, _range, relative_path, self, header=header)


class Worksheet(object):
    def __init__(self, wb, wb_formula, sheet, named_range, link, xlplugin, header=True):
        self.link = link
        self.pp = xlplugin.pp
        if named_range == "":
            if isinstance(sheet, int):
                ws = wb.sheets[sheet]
                ws_formula = wb_formula.sheets[sheet]
            else:
                ws = wb.get_sheet_by_name(sheet)
                ws_formula = wb_formula.get_sheet_by_name(sheet)
            range_str = None
            r0, c0, r1, c1 = 1, 1, ws.max_row, ws.max_column
            self.title = ws.title
        else:
            try:
                sheet, range_str = wb.get_named_range(named_range).destinations.next()
            except KeyError:
                raise Exception("Range '%s' not found"%named_range)
            if sheet is None:
                raise Exception("openpyxl 2.4.0 doesn't handle spaces in sheet names well, check that you have no spaces")
            ws = wb.get_sheet_by_name(sheet)
            ws_formula = wb_formula.get_sheet_by_name(sheet)
            c0, r0, c1, r1 = [x for x in openpyxl.utils.range_boundaries(range_str)]
            self.title = named_range

        self.width = c1 - c0 + 1
        self.height = r1 - r0 + 1

        # Remember the original openpyxl worksheet object, this is for debug only.
        self.ws = ws
        self.ws_formula = ws_formula

        # Make our grid
        self.grid = []
        for i in range(self.height):
            self.grid.append([None] * self.width)

        # Populate cells
        comment_id = 0
        for r, row in enumerate(ws.iter_rows(range_string=range_str)):
            if (r > 1000):
                error("Worksheet '%s' has more than 1000 rows!" % ws.title)

            # Calculate size for this row
            row_width = self.calc_row_width(row)

            for c, cell in enumerate(row):
                if (cell.comment != None):
                    # Increment the global table comment ID for footnotes
                    comment_id += 1
                    # Calculate the global comment ID
                    xlplugin.comment_gid += 1
                self.grid[r][c] = Cell(wb, cell, r==0 and header, row_width, comment_id, xlplugin.comment_gid)

        # Populate formulas
        for r, row in enumerate(ws_formula.iter_rows(range_string=range_str)):
            for c, cell in enumerate(row):
                self.grid[r][c].formula = self.pp.cleanstr(cell.value)

        # Populate conditional formatting
        self.apply_formatting()

        # Handle Merging
        for merge in ws.merged_cell_ranges:
            colnum, rownum, endcolnum, endrownum = [x for x in openpyxl.utils.range_boundaries(merge)]
            ri, ci = max(rownum - r0, 0), max(colnum - c0, 0)
            rj, cj = min(endrownum - r0, self.height - 1), min(endcolnum - c0, self.width - 1)
            if ri >= self.height or ci >= self.width or  rj < 0 or cj < 0: continue
            merge_cell = self.grid[ri][ci]
            merge_cell.rowspan = rj - ri + 1
            merge_cell.colspan = cj - ci + 1

            # walk merged out cells and remove them (mark as None)
            for r in range(ri, rj + 1):
                for c in range(ci, cj + 1):
                    if (r, c) != (ri, ci):
                        self.grid[r][c] = None

    def calc_row_width(self, row):
        """
        Estimate width of each element in this row based on the number
        of pixels per cell
        """
        row_width = 0
        for c, cell in enumerate(row):
            if (cell.value == None): continue
            row_width = max(self.pixelwidth(cell), row_width)
        return row_width

    def pixelwidth(self, cell):
        """
        rough approximation of pixels based on typical character widths
        """
        if (not isinstance(cell.value, basestring)):
            s = str(cell.value)
        else:
            s = cell.value

        if (cell.alignment.textRotation == 0):
            # normally aligned, just return zero
            return 0

        max_size = 0
        for line in s.split("\n"):
            size = 0
            line = line.encode('ascii', 'ignore')
            total_chars = len(line)
            line = re.sub("[A-Z_02-9ogqp*\"]", "", line)
            size += (total_chars - len(line)) * 9.0
            total_chars = len(line)
            line = re.sub("[flirt']", "", line)
            size += (total_chars - len(line)) * 5
            total_chars = len(line)
            line = re.sub("[a-z1\line]", "", line)
            size += (total_chars - len(line)) * 8.0
            size += 8 * len(line)
            max_size = max(size, max_size)
        return max_size


    def apply_formatting(self):
        """
        Apply global formatting in the worksheet object.  This is not an exhaustive implementation,
        features are added on an as-needed basis.
        """
        # Data Bars
        self.apply_data_bars(self.ws.conditional_formatting.cf_rules)

    def get_range(self, cell_range):
        """
        Convert a cell range to an x and y range

        Row is index 0, col is index 1
        """
        start, end = cell_range.split(":")

        xy = openpyxl.utils.coordinate_from_string(start)
        start_row = xy[1] - 1
        start_col = openpyxl.utils.column_index_from_string(xy[0]) - 1

        xy = openpyxl.utils.coordinate_from_string(end)
        end_row = xy[1] - 1
        end_col = openpyxl.utils.column_index_from_string(xy[0]) - 1
        return [[start_row, start_col], [end_row, end_col]]

    def apply_data_bars(self, cf_rules):
        """
        Apply data bars from conditional formatting rules
        """
        for cell_range in cf_rules.keys():
            # Look for data bars
            for rule in cf_rules[cell_range]:
                if rule.type != "dataBar":
                    continue
                # We have a data bar!
                # Simple hack here - assume we sum values across the range of cells
                # and caluclate percentage contribution per cell.  Then apply that
                # as a formatting element to our cell object.
                cell_sum = 0
                start, end = self.get_range(cell_range)
                for ri in range(start[0], end[0]+1):
                    for ci in range(start[1], end[1]+1):
                        try:
                            cell_sum += float(self.grid[ri][ci].value)
                        except:
                            # ignore this, assume zero
                            pass

                # Now, apply data bars
                for ri in range(start[0], end[0]+1):
                    for ci in range(start[1], end[1]+1):
                        try:
                            bar_pct = float(self.grid[ri][ci].value) / cell_sum
                        except:
                            # ignore this, assume zero
                            bar_pct = 0
                        self.grid[ri][ci].data_bar = bar_pct
                        self.grid[ri][ci].data_bar_rgb = rule.dataBar.color.rgb[-6:]

    def has_data(self, rows):
        """
        Seek all rows, return true if there is data
        """
        for row in rows:
            for cell in row:
                if (cell == None):
                    continue
                elif not cell.empty():
                    return True

        # no more data found
        return False

    def html(self):
        """
        Return an HTML string to cover this table
        """
        s = '<div class="table" id="%s"><table>\n'%self.pp.title2id("table-" + self.title)

        for i in range(len(self.grid)):
            row = self.grid[i]

            # Lookahead - compress empty rows at the end of the file
            if (not self.has_data(self.grid[i:])):
                break

            s += "<tr>\n"
            for cell in row:
                if (cell == None): continue
                s += cell.html(0)
            s += "</tr>\n"
        s += "</table>\n"

        s += '<p class="table_caption"><b>%s</b> (<a href=\"%s\">source</a>)</p>\n' % (self.title, self.link)

        # Add comments
        s += self.comments()

        s += '</div>\n'

        return s

    def comments(self):
        """
        Build a string of the comments, return an empty string if no comments

        To make the summary print nice, we use a table with no borders. 
        First col is the comment ID, the 2nd col is the comment.
        """
        comment_list = []
        for i in range(len(self.grid)):
            row = self.grid[i]
            for cell in self.grid[i]:
                if (cell == None): continue
                if (cell.comment != None):
                    # Create a list of cells with comments
                    comment_list.append(cell)

        s = ""
        if len(comment_list) > 0:
            s += '\n<p class="table_notes">Notes:</p><table class="notes">\n'
            for c in comment_list:
                cid = c.comment_id
                gid = c.global_comment_id
                s += "\n<tr><td valign=\"top\"><a name=\"__C%d\"><a href=\"#__TC%d\"><b>^%d^</b></a></a></td><td><a style=\"color: inherit;\" href=\"#__TC%d\"><div>%s</div></a></td></tr>\n" % (gid, gid, cid, gid, c.comment)
            s += "</table>\n"
        return s


class Cell(object):
    def __init__(self, wb, cell, isheader=False, row_width=0, comment_id=0, global_comment_id=0):
        self.cell = cell
        self.comment_id = comment_id
        self.global_comment_id = global_comment_id
        self.comment = None
        if cell.comment != None:
            # comments always start with the user name
            self.comment = "\n".join(cell.comment.text.split("\n")[1:])
        self.wb = wb
        # width of information measured in characters.  Convert to pixels on the input
        self.row_width = row_width
        self.data_bar = None
        self.number_format = cell.number_format
        self.header = isheader
        self.valign = cell.alignment.vertical
        if (self.valign == "center"): self.valign = None
        self.halign = cell.alignment.horizontal
        if (self.halign == "left"): self.halign = None
        self.colspan = 1
        self.rowspan = 1
        self.bold = cell.font.b
        self.italics = cell.font.i
        self.font_color = self.xlcolor2html(cell.font.color, "#000000")

        if (cell.fill == None):
            self.bgcolor = None
        else:
            self.bgcolor = self.xlcolor2html(cell.fill.fgColor, "#FFFFFF")

        self.value = self.cellmd(cell.value, cell.number_format)

        # Strikethrough font
        if cell.font.strike and len(self.value.strip()) > 0:
            self.value = "~~%s~~"%self.value

        self.hyperlink = cell.hyperlink
        self.formula = ""

    def xlcolor2html(self, xlcolor, default):
        if xlcolor is None:
            return None
        if xlcolor.type == 'rgb':
            try:
                if (xlcolor.rgb == "00000000"):
                    # Ignore it, this is bogus
                    return None
                ans = "#" + xlcolor.rgb[2:].upper()
            except:
                return None
        elif xlcolor.type == 'theme':
            import colorsys
            rgb = self.wb.theme_colors[xlcolor.theme]
            i = lambda s, n: eval("1./255*0x%s"%s[n:n+2])
            h,l,s = colorsys.rgb_to_hls(i(rgb, 0), i(rgb, 2), i(rgb, 4))
            if xlcolor.tint < 0:
                l += xlcolor.tint * l
            else:
                l += xlcolor.tint * (1 - l)
            ans = "hsl(%d, %d%%, %d%%)"%(360 * h, s * 100, l * 100)
        else: # unrecognized color type
            return None
        if ans == default:
            return None
        return ans

    def cellmd(self, s, fmt=None):
        """
        take an xls cell and make it markdown friendly
        """

        def escape(s):
            '''Replace special characters "&", "<" and ">" to HTML-safe sequences.
            If the optional flag quote is true, the quotation mark character (")
            is also translated.'''
            if (type(s) == str):
                s = s.replace("&", "&amp;") # Must be done first!
                s = s.replace("<", "&lt;")
                s = s.replace(">", "&gt;")

            return s

        if (not isinstance(s, basestring)):
            if (s == None):
                return ""
            else:
                if (fmt != None):
                    s = self.xlfmt(s, fmt)
                else:
                    s = str(s)

        s = s.strip()
        s = s.encode('ascii', 'ignore')
        s = escape(s)
        # Turn single newlines into a paragraph.
        s = s.replace("\n", "\n\n")
        return s

    def xlfmt(self, d, fmt):
        """
        Given an int/float/long, convert to a string using excel
        formatting rules.  TODO - support dates
          General
          0%
          0
          0.0%
          0.00
          0.000
          0.00
          0.00E+00
        """
        if (fmt == "General"):
            # No special formatting
            return str(d)

        if (fmt.startswith("#")):
            # strip leading #'s
            fmt = re.sub("^.*#", "", fmt)

        if (fmt.endswith("%")):
            # Scale to percentage
            d = d * 100.0
            fmt = fmt[:-1]
            if (fmt.find(".") >= 0):
                precision = len(fmt.split(".")[1])
                printfmt = "%%.%df%%%%" % precision
            else:
                printfmt = "%d%%"
            return printfmt % d
        m = re.search("^0\.(0+)$", fmt)
        if (m != None):
            precision = len(m.group(1))
            printfmt = "%%.%df" % precision
            return printfmt % d
        m = re.search(r"^0\.(0+)E\+00$", fmt)
        if (m != None):
            precision = len(m.group(1))
            printfmt = "%%.%de" % precision
            return printfmt % d
        return str(d)

    def empty(self):
        if (self.formula == "" and self.value == ""):
            return True
        else:
            return False

    def html(self, col):
        s = ""
        style = ""
        if self.bold: style += " font-weight:bold;"
        if self.italics: style += " font-style: italic;"
        if self.bgcolor != None and self.bgcolor != "#FFFFFF": style += " background-color:%s;" % self.bgcolor
        if self.font_color != None and self.font_color != "#000000": style += " color:%s;" % self.font_color
        if self.data_bar != None: style += " background-repeat: no-repeat; background-size: %2d%% 100%%; background-image:linear-gradient(to right, #%s, #%s);" % (int(self.data_bar * 100), self.data_bar_rgb, self.data_bar_rgb)
        if (self.header): s += "<th"
        else: s += "<td"
        if (self.colspan > 1):
            s += " colspan=\"%d\"" % self.colspan
        if (self.rowspan > 1):
            s += " rowspan=\"%d\"" % self.rowspan
        if (self.cell.alignment.textRotation == 90):
            # User has assigned vertical alignment, apply it.
            s += " align=\"left\" valign=\"bottom\" class=\"vertical_cell\""
            style += " height: %dpx" % (self.row_width)
        else:
            if (self.valign != None):
                s += " valign=\"%s\"" % self.valign
            if (self.halign != None):
                s += " align=\"%s\"" % self.halign

        if (style != ""):
            s += " style=\"%s\"" % style
        s += ">\n"

        # Hack to print out hyperlinks.  This is a very targeted solution, probably only works
        # with hsd query outputs
        m = re.search('=HYPERLINK\("([^"]+)",\s*"([^"]+)"\)', self.formula)
        if (m != None):
            hyperlink = m.group(1)
            text = m.group(2)
            if self.comment != None:
                # Embedded both the link to the comment and the backwards anchor
                gid = self.global_comment_id
                cid = self.comment_id
                text += "<a name=\"__TC%d\"><a href=\"#__C%d\">^%d^</a></a>" % (gid, gid, cid)
            if (self.cell.alignment.textRotation == 90):
                # Wrap it in a rotate div
                lines = len(text.split("\n"))
                text = "<div class=\"rotate\" style=\"width: %dpx; transform: translate(%dpx, 0) rotate(-90deg);\">%s</div>" % (lines*15, lines*15, text)
            
            s += "  <a href=\"%s\">%s</a>\n" % (hyperlink, text)
        else:
            if (self.hyperlink != None and self.hyperlink.target != None):
                # self.hyperlink.target == None when there's internal link in the document
                # internal excel links are useless in the output, so just skip them
                link = self.hyperlink.target
                if self.hyperlink.location != None:
                    link += "#%s" % self.hyperlink.location
                s += "  <a href=\"%s\">\n" % link

            # For long values give browser a chance to insert a break
            value = self.value
            if len(value) > 15 and self.cell.alignment.textRotation != 90:
                # cell rotation and breaks don't play nice
                value = value.replace("_", "_<wbr>")
                value = re.sub(r"(?<!\\)\[", r"<wbr>[", value)

            # Escape single dash, otherwise it will be converted to a list (dot)
            if value.strip() == '-':
                value = '&#8208;' # hyphen (shorter than ndash)

            # For a string of dashes or equals, escape them so as to not mess up markdown
            # conversion
            if re.search("^[-=]+$", value.strip()) != None:
                value = value.replace("-", "&#8208;")
                value = value.replace("=", "&#61;")

            if self.comment != None:
                # Embedded both the link to the comment and the backwards anchor
                gid = self.global_comment_id
                cid = self.comment_id
                value += "<a name=\"__TC%d\"><a href=\"#__C%d\">^%d^</a></a>" % (gid, gid, cid)

            if (self.cell.alignment.textRotation == 90):
                # Wrap it in a rotate div
                lines = len(value.split("\n"))
                # Setting up a box that is 15px wide per line (roughly the size of the pixels in this style).
                # TODO: javascript to scale this based on font size?
                # Rotate left by 90deg pivot at the bottom left
                # Shift right by the width to place it back inside the table cell
                s += "<div class=\"rotate\" style=\"width: %dpx; transform: translate(%dpx, 0) rotate(-90deg);\">%s</div>" % (lines*15, lines*15, value)
            else:
                s += "  %s\n" % value
            if (self.hyperlink != None):
                s += "  </a>\n"
        if (self.header): s += "</th>"
        else: s += "</td>"

        return self.indent(s, col)

    def indent(self, lines, column):
        """
        Insert spaces to shift this string right by the column count
        """
        columns = " " * column
        output_string = ""

        if (not isinstance(lines, list)):
            lines = lines.splitlines()

        for line in lines:
            output_string += columns + line + "\n"

        return output_string


new = XlsPlugin
